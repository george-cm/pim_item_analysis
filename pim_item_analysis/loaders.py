"""Load Excel file into a list of lists.

Parses an Excel file at the given file path and returns
a list containing the rows and columns of the first worksheet.
Each inner list represents a row, with the cell values as elements.
"""

import csv
import datetime
import io
import sqlite3
from pathlib import Path
from typing import Any
from typing import Dict
from typing import Iterator
from typing import List
from typing import Optional
from typing import Tuple

import openpyxl

from pim_item_analysis.db import db_add_label
from pim_item_analysis.db import db_create_table
from pim_item_analysis.db import db_drop_tables
from pim_item_analysis.db import file_prefix
from pim_item_analysis.db import get_export_date_from_file
from pim_item_analysis.db import normalize_name
from pim_item_analysis.db import round_seconds


def load_pimfile_to_db(
    conn: sqlite3.Connection,
    file_path: Path,
    drop_table_first: bool = False,
    unique_index_columns: Optional[List[str]] = None,
    label: str | None = None,
) -> int:
    """Load file into the database"""

    header_maps: Dict[str, Dict[str, str]] = {
        "item_texts": {
            "Item.Item no.": "Item no.",
            "SKU": "SKU",
            "Language-specific data.Language": "Language",
            "Language-specific data.Item Name": "Item Name",
            "Language-specific data.Item Short Description": "Item Short Description",
            "Language-specific data.Item Long description": "Item Long description",
        }
    }

    current_file_suffix: str = file_prefix(file_path)
    export_date: datetime.datetime = get_export_date_from_file(file_path)
    print(current_file_suffix)
    print(export_date)
    with file_path.open(encoding="utf-8") as f:
        csv_reader: Iterator[List[str]] = csv.reader(f)
        header: List[str] = next(csv_reader)
        if current_file_suffix in header_maps:
            header = [header_maps[current_file_suffix][x] for x in header]

        columns: list[str] = [x for x in (["export_date"] + header)]
        columns_str: str = ", ".join([f"[{x}]" for x in columns])
        extra_fields: dict[str, str] = {
            "id": "INTEGER PRIMARY KEY AUTOINCREMENT",
            "export_date": "DATETIME NOT NULL",
        }
        fields: dict[str, str] = {**extra_fields, **{k: "TEXT" for k in columns[1:]}}
        if drop_table_first:
            db_drop_tables(conn, [current_file_suffix])
        db_create_table(
            conn, current_file_suffix, fields, unique_index_columns=unique_index_columns
        )

        sql: str = f"""
            INSERT OR IGNORE INTO {current_file_suffix} ({columns_str})
            VALUES ({','.join(['?' for _ in columns])})
        """
        data: List[List[str | datetime.datetime]] = [
            [export_date] + row for row in csv_reader
        ]

        cursor = conn.cursor()
        cursor.executemany(sql, data)
        inserted_row_count = cursor.rowcount
        conn.commit()

        # add the label
        if label and inserted_row_count > 0:
            db_add_label(
                conn, dataset_type="pim", dataset_datetime=export_date, label=label
            )
    return inserted_row_count


def load_hybris_excel_to_db(
    conn: sqlite3.Connection,
    file_path: Path,
    drop_table_first: bool = False,
    unique_index_columns: Optional[List[str]] = None,
    label: str | None = None,
) -> int:
    """Load Excel file to database."""
    current_file_suffix: str = file_prefix(file_path)
    export_date: datetime.datetime = round_seconds(
        datetime.datetime.fromtimestamp(file_path.stat().st_ctime)
    )
    # print(export_date)
    with file_path.open("rb") as f:
        in_memory_file = io.BytesIO(f.read())

    workbook: openpyxl.Workbook = openpyxl.load_workbook(in_memory_file, read_only=True)
    sheet = workbook.worksheets[0]

    row_iter = sheet.iter_rows(values_only=True)  # type: ignore
    header: tuple[str] = next(row_iter)  # type: ignore
    columns = list(("export_date", "Item no.", *header))
    columns_str = ", ".join([f"[{x}]" for x in columns])
    extra_fields: dict[str, str] = {
        "id": "INTEGER PRIMARY KEY AUTOINCREMENT",
        "export_date": "DATETIME NOT NULL",
    }
    fields: dict[str, str] = {**extra_fields, **{k: "TEXT" for k in columns[1:]}}

    if drop_table_first:
        db_drop_tables(conn, [current_file_suffix])

    db_create_table(
        conn, current_file_suffix, fields, unique_index_columns=unique_index_columns
    )

    data: List[List[Any]] = [
        [export_date, f"{row[1]}~{row[0]}", *row] for row in row_iter
    ]

    sql: str = f"""
        INSERT OR IGNORE INTO {current_file_suffix} ({columns_str})
        VALUES ({','.join(['?' for _ in columns])})
    """

    cursor: sqlite3.Cursor = conn.cursor()
    cursor.executemany(sql, data)
    inserted_row_count: int = cursor.rowcount
    conn.commit()

    # add the label
    if label and inserted_row_count > 0:
        db_add_label(
            conn, dataset_type="hybris", dataset_datetime=export_date, label=label
        )
    return inserted_row_count


def load_docfile_into_db(
    conn: sqlite3.Connection,
    file_path: Path,
    prefix: str,
    request_date: datetime.datetime,
    config: Dict[str, Any],
    drop_table_first: bool = False,
    unique_index_columns: Optional[List[str]] = None,
    label: str | None = None,
) -> int:
    """Load doc Excel xlsx file into the database"""
    total_inserted_row_count: int = 0
    cursor: sqlite3.Cursor = conn.cursor()
    with file_path.open("rb") as f:
        in_memory_file = io.BytesIO(f.read())

    sheets: List[str] = config["doc"]["file_sheets"][prefix]
    workbook: openpyxl.Workbook = openpyxl.load_workbook(
        in_memory_file, read_only=True, data_only=True
    )
    for sh in workbook.worksheets:
        if sh.title.lower() not in sheets:
            continue
        # print(f"sheet name: {sh.title}")
        table_name: str = normalize_name(sh.title)
        config_header = list(config["doc"]["sheet_headers"][sh.title.lower()].keys())
        # config_header_requirements = list(
        #     config["doc"]["sheet_headers"][sh.title.lower()].values()
        # )
        # print(f"{config_header=}")
        # print(f"{config_header_requirements=}")
        start_row_header = config["doc"]["start_rows"][sh.title.lower()]["header"]
        start_row_data = config["doc"]["start_rows"][sh.title.lower()]["data"]
        data: List[Any] = []
        empty_rows: int = 0
        for i, row in enumerate(sh.iter_rows(values_only=True), start=1):  # type: ignore
            # getting rid of extra columns from the end which are not
            # parte of the inital template that the user might have added
            trimmed_row: Tuple[str | float | datetime.datetime | None, ...] = row[
                : len(config_header)
            ]
            if i == start_row_header:
                columns: List[str] = list(("request_date", *trimmed_row, "file_name"))  # type: ignore
                columns_str: str = ", ".join([f"[{x}]" for x in columns])
                extra_fields: dict[str, str] = {
                    "id": "INTEGER PRIMARY KEY AUTOINCREMENT",
                    "request_date": "DATETIME NOT NULL",
                }
                fields: dict[str, str] = {
                    **extra_fields,
                    **{
                        k: "DATE" if "date" in k.lower() else "TEXT"
                        for k in columns[1:]
                    },
                }
                if drop_table_first:
                    db_drop_tables(conn, [table_name])
                db_create_table(
                    conn,
                    table_name,
                    fields,
                    unique_index_columns=unique_index_columns,
                )
                sql: str = f"""
                    INSERT OR IGNORE INTO {table_name} ({columns_str})
                    VALUES ({','.join(['?' for _ in columns])})
                """
                # print(f"{i}:header: {row}")
            if i >= start_row_data:
                # print(f"{i}:data: {row}")
                data_row: List[Any] = [
                    x.strip()
                    if isinstance(x, str)
                    else (x.date() if isinstance(x, datetime.datetime) else x)
                    for x in trimmed_row
                ]
                # print(f"{data_row=}")
                if any(data_row):
                    data.append([request_date] + data_row + [file_path.name])
                else:
                    empty_rows += 1
            if empty_rows >= 3:
                break
        cursor.executemany(sql, data)
        inserted_row_count: int = cursor.rowcount
        total_inserted_row_count += inserted_row_count
        conn.commit()

        print(
            f"Inserted {inserted_row_count} rows from sheet {sh.title} from file {file_path}\n"
        )
    # add the label
    if label and inserted_row_count > 0:
        db_add_label(
            conn, dataset_type="doc", dataset_datetime=request_date, label=label
        )
    return total_inserted_row_count


def main():
    """Main function."""
    print("This module doesn't do anything on it's own.")


if __name__ == "__main__":
    main()
